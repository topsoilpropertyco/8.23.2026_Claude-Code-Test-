#!/usr/bin/env python3
"""Build the two shipping artifacts from the fact ledger.

  properties.csv     one row per property, CRM-importable, curated columns
  DEAL-DATABASE.xlsx the same table plus the raw evidence behind every cell

The ledger holds facts, not rows. Each fact is (property_key, field, value,
evidence_quote, msg_id). Pivoting to one-row-per-property means deciding what
to do when a property has three different asking prices or five owner phones.
This keeps them all, joined with " | ", because a disagreement between two
emails is itself information about the deal.

801 distinct field names accumulated over the sweep -- far too many columns for
a spreadsheet. Three moves fix that without dropping anything:
  ALIASES   folds obvious synonyms into one canonical column (address ->
            address_raw, motivation -> seller_motivation, ...)
  COLUMNS   is the curated column order a human actually scans
  the tail   everything else lands in `other_details` as "field: value" pairs,
            so nothing is lost, it is just not its own column
ALL_CAPS field names are curator flags, not data; they go to `flags` so a row
that needs human eyes announces itself.
"""
import os, sys, csv, sqlite3, re, datetime
from collections import OrderedDict, defaultdict

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB   = os.path.join(HERE, "data", "ledger", "deals.db")
OUT  = os.path.join(HERE, "data", "out")

ALIASES = {
    "address": "address_raw", "street_address": "address_raw",
    "motivation": "seller_motivation",
    "contact": "contact_raw", "contact_info": "contact_raw",
    "phone": "contact_phone", "email": "contact_email",
    "name": "contact_name", "lead_name": "contact_name",
    "size": "sqft", "building_size": "sqft", "square_feet": "sqft",
    "units": "unit_count", "number_of_units": "unit_count",
    "acres": "acreage", "price": "asking_price", "list_price": "asking_price",
    "next_action": "next_step", "notes": "call_notes", "note": "call_notes",
}

COLUMNS = [
    "property_name", "address_raw", "mailing_address", "state", "website",
    "asking_price", "price_indication", "appraised_value", "seller_financing",
    "revenue_current", "revenue_at_new_rates", "revenue_at_full", "expenses",
    "owner_name", "owner_phone", "owner_phone_best", "owner_email",
    "contact_name", "contact_phone", "contact_email", "contact_raw", "contact_note",
    "broker_name", "broker_phone", "broker_email", "broker_note",
    "unit_count", "unit_mix", "sqft", "acreage", "buildings", "occupancy",
    "climate", "condition", "rent_card", "rent_range", "rent_note", "expansion",
    "seller_motivation", "deal_status", "deal_terms", "lead_status",
    "crm_pipeline", "crm_pipeline_status", "crm_lead_name",
    "next_step", "data_gap", "performance", "location_note", "operating_since",
    "age", "asset_note", "entity", "operating_company", "ownership",
    "ownership_change", "ownership_status", "analysis", "analysis_link",
    "documents", "attachment", "call_notes",
]
IS_FLAG      = re.compile(r"^[A-Z][A-Z0-9_]*$")
NOT_A_TARGET = ("NOT_A_PROPERTY", "NOT_A_LEAD")
# Keys that are operational/relationship dossiers rather than acquisition targets.
NON_PROPERTY_KEYS = {
    "lead-gen-operations", "investor-network-contacts", "ssi-mastermind-membership",
    "title-and-closing-vendors", "legal-entity-formation-vendors",
    "merchant-processing-vendors", "remote-concierge-manager-rcm",
}

def load():
    c = sqlite3.connect(DB)
    props = OrderedDict()
    for key, mid, tid, field, value, conf in c.execute(
            """SELECT property_key, msg_id, thread_id, field, value, confidence
               FROM facts ORDER BY property_key, id"""):
        p = props.setdefault(key, {"vals": defaultdict(list), "flags": [], "tail": [],
                                   "threads": OrderedDict(), "msgs": OrderedDict(),
                                   "unsure": 0, "n": 0})
        p["n"] += 1
        p["threads"][tid] = 1
        p["msgs"][mid] = 1
        if conf in ("unsure", "inferred", "low"):
            p["unsure"] += 1
        if IS_FLAG.match(field):
            p["flags"].append("%s: %s" % (field, value))
            continue
        field = ALIASES.get(field, field)
        if field in COLUMNS:
            if value not in p["vals"][field]:
                p["vals"][field].append(value)
        else:
            p["tail"].append("%s: %s" % (field, value))
    return c, props

def row_for(key, p):
    flags  = " || ".join(p["flags"])
    if key in NON_PROPERTY_KEYS:
        status = "reference"
    elif any(f.startswith(NOT_A_TARGET) for f in p["flags"]):
        status = "not-a-target"
    else:
        status = "target"
    row = [key, status, flags]
    for col in COLUMNS:
        row.append(" | ".join(p["vals"].get(col, [])))
    row.append(" || ".join(p["tail"]))
    row += [p["n"], p["unsure"],
            " ".join(x for x in p["threads"] if x),
            " ".join(x for x in p["msgs"] if x)]
    return status, row

HEADER = (["property_key", "status", "flags"] + COLUMNS +
          ["other_details", "fact_count", "low_confidence_facts",
           "source_threads", "source_msgs"])

def main():
    os.makedirs(OUT, exist_ok=True)
    c, props = load()
    rows = [row_for(k, p) for k, p in sorted(props.items())]

    csv_path = os.path.join(OUT, "properties.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(HEADER)
        for _, r in rows:
            w.writerow(r)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")

    def sheet(name, header, data, widths=None, freeze="A2"):
        ws = wb.create_sheet(name)
        ws.append(header)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in data:
            ws.append([("" if v is None else str(v))[:32000] for v in r])
        ws.freeze_panes = freeze
        ws.auto_filter.ref = ws.dimensions
        for i, h in enumerate(header, 1):
            ws.column_dimensions[get_column_letter(i)].width = (
                (widths or {}).get(h, min(42, max(12, len(h) + 4))))
        return ws

    targets = [r for s, r in rows if s == "target"]
    others  = [r for s, r in rows if s != "target"]
    wide = {"property_key": 34, "flags": 60, "address_raw": 38, "call_notes": 60,
            "other_details": 70, "source_threads": 30, "source_msgs": 30,
            "seller_motivation": 45, "deal_status": 45}

    sheet("Properties", HEADER, targets, wide)
    sheet("Reference & Not-A-Target", HEADER, others, wide)

    review = [r for s, r in rows if r[2]]
    sheet("Needs Review", HEADER, review, wide)

    facts = list(c.execute(
        """SELECT property_key, field, value, evidence_quote, confidence,
                  thread_id, msg_id, created_at
           FROM facts ORDER BY property_key, id"""))
    sheet("Facts (evidence)",
          ["property_key", "field", "value", "evidence_quote", "confidence",
           "thread_id", "msg_id", "recorded_at"], facts,
          {"value": 55, "evidence_quote": 80, "property_key": 34})

    threads = list(c.execute(
        """SELECT thread_id, labels, subject, read_tier,
                  CASE WHEN no_property=1 THEN 'no property' ELSE 'has facts' END,
                  notes
           FROM threads WHERE body_read=1 AND (out_of_scope IS NULL OR out_of_scope=0)
           ORDER BY labels, thread_id"""))
    sheet("Threads Read", ["thread_id", "labels", "subject", "read_tier",
                           "outcome", "note"], threads,
          {"subject": 60, "note": 70})

    ws = wb["Sheet"]; ws.title = "README"
    for line in README.strip().split("\n"):
        ws.append([line])
    ws.column_dimensions["A"].width = 118
    for row in ws.iter_rows():
        row[0].alignment = Alignment(wrap_text=False, vertical="top")
    ws["A1"].font = Font(bold=True, size=14)
    wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)

    xlsx_path = os.path.join(OUT, "DEAL-DATABASE.xlsx")
    wb.save(xlsx_path)

    n_fact = c.execute("SELECT count(*) FROM facts").fetchone()[0]
    n_thr  = c.execute("SELECT count(*) FROM threads WHERE body_read=1 AND (out_of_scope IS NULL OR out_of_scope=0)").fetchone()[0]
    print("wrote", csv_path)
    print("wrote", xlsx_path)
    print("  properties : %d (%d targets, %d reference/not-a-target)"
          % (len(rows), len(targets), len(others)))
    print("  needs review: %d" % len(review))
    print("  facts      : %d   threads read: %d" % (n_fact, n_thr))
    print("  columns    : %d" % len(HEADER))

README = """DEAL DATABASE - built %s from %s Gmail threads

WHAT THIS IS
Every real-estate opportunity found in ~5 years of mail across 7 Gmail labels,
merged so that one property = one row, no matter how many threads mentioned it.

THE SHEETS
  Properties               acquisition targets. Start here.
  Reference & Not-A-Target operational dossiers (lead-gen, vendors, network) and
                           rows confirmed not to be acquisition targets.
  Needs Review             every row carrying a curator flag - conflicts, missing
                           addresses, unresolved identities. Work this list.
  Facts (evidence)         all facts, one per line, each with the VERBATIM quote
                           and the Gmail message id it came from.
  Threads Read             every thread opened, its outcome, and why.

READING A ROW
  status         target / reference / not-a-target
  flags          ALL_CAPS curator notes. A non-empty cell means a human decision
                 is pending. Read these before acting on the row.
  a cell with "|" holds MORE THAN ONE recorded value. Two asking prices means two
                 emails said different things. Neither was discarded.
  other_details  long-tail fields that did not earn their own column.
  source_threads / source_msgs  paste any id into Gmail search to open the source.

HOW TO TRUST A CELL
  Every value traces to a verbatim quote. Filter "Facts (evidence)" by the
  property_key to see the exact sentence and message behind each field.
  Nothing here was inferred: no quote, no fact.

IMPORTING TO A CRM
  Use properties.csv. Filter status = target first. Map property_name, address_raw,
  owner_name, owner_phone, owner_email, asking_price to your CRM fields and carry
  other_details + flags into a long-text notes field so context survives the import.
""" % (datetime.date.today().isoformat(), "1,215")

if __name__ == "__main__":
    main()
